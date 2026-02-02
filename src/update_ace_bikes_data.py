import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os

# Set random seed for reproducibility
np.random.seed(42)

# ============================================================================
# 1. LOAD DATA
# ============================================================================

print("Loading data files...")

# Load main Excel file
df_main = pd.read_excel('./data_original/Ace_Bikes_Data.xlsx', sheet_name='Data')

# Load new employees
df_new_employees = pd.read_excel('./data_new/newEmployees.xlsx')

# Load new line item sales
df_new_line_items = pd.read_csv('./data_new/newLineItemSales.csv')

print(f"Loaded {len(df_new_employees)} new employees")
print(f"Loaded {len(df_new_line_items)} new line items")

# ============================================================================
# 2. IDENTIFY TABLE BOUNDARIES IN THE MAIN SHEET
# ============================================================================

print("\nIdentifying table boundaries...")

# Find where each table starts based on column headers
# EmployeeDates: columns 0-3 (EmployeeID, StartDate, TerminationDate, LocationID)
# EmployeeReviewInfo: columns 5-11 (EmpID, Date, Salesmanship, ProductKnowledge, TeamPlayer, Innovator, Satisfaction)
# EmployeeTerminationReasons: columns 13-14 (EmployeeID, Reason)
# LineItemReturns: columns 40-41 (LineItemID, ReturnID)

# Find the last row with data for each table
employee_dates_last_row = df_main['EmployeeID'].notna().sum() - 1
employee_review_last_row = df_main['EmpID'].notna().sum() - 1
employee_termination_last_row = df_main['EmployeeID.1'].notna().sum() - 1
line_item_returns_last_row = df_main['LineItemID'].notna().sum() - 1

print(f"EmployeeDates ends at row: {employee_dates_last_row}")
print(f"EmployeeReviewInfo ends at row: {employee_review_last_row}")
print(f"EmployeeTerminationReasons ends at row: {employee_termination_last_row}")
print(f"LineItemReturns ends at row: {line_item_returns_last_row}")

# ============================================================================
# 3. APPEND NEW EMPLOYEE DATES TO EMPLOYEEDATES TABLE
# ============================================================================

print("\n" + "="*80)
print("TASK 1: Appending new employee dates...")
print("="*80)

# Extract only the required columns
new_employee_dates = df_new_employees[['EmployeeID', 'StartDate', 'TerminationDate', 'LocationID']].copy()

print(f"\nAppending {len(new_employee_dates)} employee records to EmployeeDates table")

# Add new employee dates starting from the next available row
start_row = employee_dates_last_row + 1
for idx, row in new_employee_dates.iterrows():
    df_main.loc[start_row + idx, 'EmployeeID'] = row['EmployeeID']
    df_main.loc[start_row + idx, 'StartDate'] = row['StartDate']
    df_main.loc[start_row + idx, 'TerminationDate'] = row['TerminationDate']
    df_main.loc[start_row + idx, 'LocationID'] = row['LocationID']

print(f"✓ Added {len(new_employee_dates)} employee date records")

# ============================================================================
# 4. GENERATE EMPLOYEE REVIEW INFO FOR NEW EMPLOYEES
# ============================================================================

print("\n" + "="*80)
print("TASK 2: Generating employee skill reviews...")
print("="*80)

def generate_review_dates(start_date, end_date=None):
    """Generate semi-annual review dates (January and July)"""
    if pd.isna(start_date):
        return []
    
    start = pd.to_datetime(start_date)
    
    # If no end date, use current date (2023-12-31 as a reasonable end)
    if pd.isna(end_date):
        end = pd.to_datetime('2023-12-31')
    else:
        end = pd.to_datetime(end_date)
    
    review_dates = []
    
    # Start from the first review period after start date
    # Reviews are in January (around 10th) and July (around 2nd-8th)
    current_year = start.year
    
    # Check if started before July, add July review
    if start <= pd.to_datetime(f'{current_year}-07-01'):
        july_review = pd.to_datetime(f'{current_year}-07-{np.random.randint(2, 9)}')
        if july_review >= start and july_review <= end:
            review_dates.append(july_review)
    
    # Move to next January
    current_year += 1
    
    while True:
        # January review
        jan_review = pd.to_datetime(f'{current_year}-01-{np.random.randint(9, 17)}')
        if jan_review <= end:
            review_dates.append(jan_review)
        else:
            break
        
        # July review
        july_review = pd.to_datetime(f'{current_year}-07-{np.random.randint(1, 9)}')
        if july_review <= end:
            review_dates.append(july_review)
        else:
            break
        
        current_year += 1
    
    return review_dates


def generate_rating(has_training, base_mean=3.5, base_std=0.8):
    """
    Generate a rating score (1-5).
    If employee has training, boost the mean rating.
    """
    if has_training:
        # Training increases mean by 0.5-0.8 points
        mean = base_mean + np.random.uniform(0.5, 0.8)
    else:
        mean = base_mean
    
    # Generate rating with some randomness
    rating = np.random.normal(mean, base_std)
    
    # Clip to valid range and round to 1 decimal
    rating = np.clip(rating, 2.0, 5.0)
    return round(rating, 1)


# Generate reviews for each new employee
new_reviews = []

for _, emp in df_new_employees.iterrows():
    emp_id = emp['EmployeeID']
    start_date = emp['StartDate']
    termination_date = emp['TerminationDate']
    
    # Get training flags
    skills_training = emp['SkillsTraining']
    salesmanship_training = emp['SalesmanshipTraining']
    product_training = emp['ProductTraining']
    
    # Generate review dates
    review_dates = generate_review_dates(start_date, termination_date)
    
    # Generate reviews for each date
    for review_date in review_dates:
        review = {
            'EmpID': emp_id,
            'Date': review_date,
            'Salesmanship': generate_rating(salesmanship_training, base_mean=3.4),
            'ProductKnowledge': generate_rating(product_training, base_mean=3.6),
            'TeamPlayer': generate_rating(skills_training, base_mean=3.7),
            'Innovator': generate_rating(skills_training, base_mean=3.5),
            'Satisfaction': generate_rating(skills_training, base_mean=3.3)
        }
        new_reviews.append(review)

df_new_reviews = pd.DataFrame(new_reviews)

print(f"\nGenerated {len(df_new_reviews)} review records for {len(df_new_employees)} employees")
print(f"Average reviews per employee: {len(df_new_reviews) / len(df_new_employees):.1f}")

# Sample of generated reviews
print("\nSample of generated reviews:")
print(df_new_reviews.head(10))

# Append to EmployeeReviewInfo table
start_row = employee_review_last_row + 1
for idx, row in df_new_reviews.iterrows():
    df_main.loc[start_row + idx, 'EmpID'] = row['EmpID']
    df_main.loc[start_row + idx, 'Date'] = row['Date']
    df_main.loc[start_row + idx, 'Salesmanship'] = row['Salesmanship']
    df_main.loc[start_row + idx, 'ProductKnowledge'] = row['ProductKnowledge']
    df_main.loc[start_row + idx, 'TeamPlayer'] = row['TeamPlayer']
    df_main.loc[start_row + idx, 'Innovator'] = row['Innovator']
    df_main.loc[start_row + idx, 'Satisfaction'] = row['Satisfaction']

print(f"✓ Added {len(df_new_reviews)} employee review records")

# ============================================================================
# 5. ADD TERMINATION REASONS (IF ANY TERMINATED EMPLOYEES)
# ============================================================================

print("\n" + "="*80)
print("TASK 3: Adding termination reasons...")
print("="*80)

# Find employees with termination dates
terminated_employees = df_new_employees[df_new_employees['TerminationDate'].notna()]
non_terminated_employees = df_new_employees[df_new_employees['TerminationDate'].isna()]

termination_records = []

if len(terminated_employees) > 0:
    print(f"\nFound {len(terminated_employees)} terminated employees")
    
    # Define termination reasons with probabilities
    reasons = ['Another Job', 'Moved', 'Terminated']
    probabilities = [0.50, 0.25, 0.25]
    
    # Generate random reasons for terminated employees
    for _, emp in terminated_employees.iterrows():
        reason = np.random.choice(reasons, p=probabilities)
        termination_records.append({
            'EmployeeID': emp['EmployeeID'],
            'Reason': reason
        })
    
    print(f"  ✓ Assigned termination reasons to {len(terminated_employees)} employees")
else:
    print("\nNo terminated employees found")

# Add 'null' for non-terminated employees
if len(non_terminated_employees) > 0:
    print(f"\nFound {len(non_terminated_employees)} non-terminated employees")
    
    for _, emp in non_terminated_employees.iterrows():
        termination_records.append({
            'EmployeeID': emp['EmployeeID'],
            'Reason': 'null'
        })
    
    print(f"  ✓ Added 'null' reason for {len(non_terminated_employees)} active employees")

# Create dataframe with all termination records
df_terminations = pd.DataFrame(termination_records)

# Append to EmployeeTerminationReasons table
start_row = employee_termination_last_row + 1
for idx, row in df_terminations.iterrows():
    df_main.loc[start_row + idx, 'EmployeeID.1'] = row['EmployeeID']
    df_main.loc[start_row + idx, 'Reason'] = row['Reason']

print(f"\n✓ Added {len(df_terminations)} total termination reason records")

if len(terminated_employees) > 0:
    print("\nTermination reasons breakdown:")
    print(df_terminations['Reason'].value_counts())
else:
    print(f"  All {len(df_terminations)} employees have 'null' reason (active employees)")

# ============================================================================
# 6. GENERATE LINE ITEM RETURNS
# ============================================================================

print("\n" + "="*80)
print("TASK 4: Generating line item returns...")
print("="*80)

# Sample 4.5% to 5% of line items
sample_rate = np.random.uniform(0.045, 0.050)
n_returns = int(len(df_new_line_items) * sample_rate)

print(f"\nSampling {sample_rate:.2%} of {len(df_new_line_items)} line items = {n_returns} returns")

# Randomly sample line items
sampled_line_items = df_new_line_items.sample(n=n_returns, random_state=42)

# Assign ReturnIDs with specified probabilities
return_ids = ['R1', 'R2', 'R3']
return_probabilities = [0.34, 0.34, 0.32]

returns_data = []
for _, item in sampled_line_items.iterrows():
    return_id = np.random.choice(return_ids, p=return_probabilities)
    returns_data.append({
        'LineItemID': item['LineItemID'],
        'ReturnID': return_id
    })

df_returns = pd.DataFrame(returns_data)

print(f"\nReturn ID distribution:")
print(df_returns['ReturnID'].value_counts().sort_index())
print(f"\nPercentages:")
print(df_returns['ReturnID'].value_counts(normalize=True).sort_index())

# Append to LineItemReturns table
start_row = line_item_returns_last_row + 1
for idx, row in df_returns.iterrows():
    df_main.loc[start_row + idx, 'LineItemID'] = row['LineItemID']
    df_main.loc[start_row + idx, 'ReturnID'] = row['ReturnID']

print(f"✓ Added {len(df_returns)} line item return records")

# ============================================================================
# 7. CLEAN UP DATA BEFORE SAVING
# ============================================================================

print("\n" + "="*80)
print("CLEANING UP DATA...")
print("="*80)

# Convert datetime columns to date-only format
print("\nConverting datetime columns to date-only format...")

# List of date columns to convert
date_columns = ['StartDate', 'TerminationDate', 'Date', 'Date.1']

for col in date_columns:
    if col in df_main.columns:
        # For datetime columns, normalize to remove time component
        df_main[col] = pd.to_datetime(df_main[col], errors='coerce')
        # Normalize to midnight (removes time component)
        df_main[col] = df_main[col].dt.normalize()
        print(f"  ✓ Normalized {col} to date-only format")

# Rename columns that start with "Unnamed"
print("\nRenaming 'Unnamed' columns...")

unnamed_cols = [col for col in df_main.columns if str(col).startswith('Unnamed')]
if unnamed_cols:
    # Create a mapping of old names to new names (empty strings)
    rename_dict = {col: '' for col in unnamed_cols}
    
    # However, pandas doesn't allow duplicate empty column names
    # So we'll use a placeholder approach
    for i, col in enumerate(unnamed_cols):
        rename_dict[col] = f'__BLANK_{i}__'
    
    df_main.rename(columns=rename_dict, inplace=True)
    print(f"  ✓ Renamed {len(unnamed_cols)} 'Unnamed' columns to blank placeholders")
else:
    print("  ✓ No 'Unnamed' columns found")

# ============================================================================
# 8. SAVE UPDATED FILE
# ============================================================================

print("\n" + "="*80)
print("SAVING UPDATED FILE...")
print("="*80)

# Create output directory
output_dir = './data_new/'
os.makedirs(output_dir, exist_ok=True)

output_path = os.path.join(output_dir, 'newAce_Bikes_Data.xlsx')

# Save to Excel with specific date format
print(f"\nSaving to: {output_path}")

# First, save the dataframe
df_main.to_excel(output_path, sheet_name='Data', index=False)

# Then, open with openpyxl to apply formatting
from openpyxl import load_workbook
from openpyxl.styles import numbers

print("Applying date formatting to Excel...")

wb = load_workbook(output_path)
ws = wb['Data']

# Clean up column headers - replace __BLANK_X__ with None
for cell in ws[1]:
    if cell.value and '__BLANK_' in str(cell.value):
        cell.value = None

# Find date columns and apply date-only formatting
date_column_names = ['StartDate', 'TerminationDate', 'Date', 'Date.1']
header_row = [cell.value for cell in ws[1]]

date_column_indices = []
for col_name in date_column_names:
    if col_name in header_row:
        col_idx = header_row.index(col_name) + 1  # openpyxl is 1-indexed
        date_column_indices.append(col_idx)

# Apply date-only format (yyyy-mm-dd) to all date columns
for col_idx in date_column_indices:
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    col_name = ws.cell(row=1, column=col_idx).value
    
    # Apply to all rows in this column (skip header)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            # Set number format to date only (no time)
            cell.number_format = 'YYYY-MM-DD'
    
    print(f"  ✓ Applied date-only format to column {col_letter} ({col_name})")

# Save the workbook
wb.save(output_path)
print(f"  ✓ Formatting applied successfully")

print("✓ File saved successfully!")

# ============================================================================
# 9. SUMMARY REPORT
# ============================================================================

print("\n" + "="*80)
print("SUMMARY REPORT")
print("="*80)

print(f"""
Employee Updates:
  • New employees added: {len(new_employee_dates)}
  • Employee reviews generated: {len(df_new_reviews)}
  • Termination reasons added: {len(terminated_employees)}

Line Item Updates:
  • Total line items: {len(df_new_line_items)}
  • Returns generated: {len(df_returns)} ({sample_rate:.2%})
  • Return distribution:
    - R1: {(df_returns['ReturnID'] == 'R1').sum()} ({(df_returns['ReturnID'] == 'R1').mean():.1%})
    - R2: {(df_returns['ReturnID'] == 'R2').sum()} ({(df_returns['ReturnID'] == 'R2').mean():.1%})
    - R3: {(df_returns['ReturnID'] == 'R3').sum()} ({(df_returns['ReturnID'] == 'R3').mean():.1%})

Output:
  • File: newAce_Bikes_Data.xlsx
  • Location: {output_dir}
""")

print("="*80)
print("ALL TASKS COMPLETED SUCCESSFULLY!")
print("="*80)